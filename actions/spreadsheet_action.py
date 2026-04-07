"""Spreadsheet action module for RabAI AutoClick.

Provides spreadsheet operations:
- SpreadsheetReadAction: Read spreadsheet data
- SpreadsheetWriteAction: Write data to spreadsheet
- SpreadsheetCreateAction: Create new spreadsheet
- SpreadsheetSheetNamesAction: Get sheet names
- SpreadsheetCellReadAction: Read specific cell
- SpreadsheetCellWriteAction: Write to specific cell
- SpreadsheetAppendRowAction: Append row of data
- SpreadsheetFilterAction: Filter rows by condition
- SpreadsheetSortAction: Sort spreadsheet
- SpreadsheetPivotAction: Create pivot table summary
"""

import os
import csv
import json
from typing import Any, Dict, List, Optional, Union

import sys
import os
_parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _parent_dir)
from core.base_action import BaseAction, ActionResult

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SpreadsheetReadAction(BaseAction):
    """Read spreadsheet data."""
    action_type = "spreadsheet_read"
    display_name = "读取电子表格"
    description = "读取Excel或CSV文件数据"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute read.

        Args:
            context: Execution context.
            params: Dict with file_path, sheet_name, output_var.

        Returns:
            ActionResult with spreadsheet data.
        """
        file_path = params.get('file_path', '')
        sheet_name = params.get('sheet_name', '')
        output_var = params.get('output_var', 'spreadsheet_data')
        max_rows = params.get('max_rows', 0)

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_sheet = context.resolve_value(sheet_name) if sheet_name else ''
            resolved_max = context.resolve_value(max_rows) if max_rows else 0

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                data = self._read_csv(resolved_path, resolved_max)
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装: pip install openpyxl"
                    )
                data = self._read_excel(resolved_path, resolved_sheet, resolved_max)
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            context.set(output_var, data)

            return ActionResult(
                success=True,
                message=f"已读取 {len(data)} 行数据",
                data={'rows': len(data), 'output_var': output_var}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"读取电子表格失败: {str(e)}"
            )

    def _read_csv(self, path: str, max_rows: int) -> List[List[str]]:
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = list(csv.reader(f))
            if max_rows > 0:
                return reader[:max_rows + 1]
            return reader

    def _read_excel(self, path: str, sheet_name: str, max_rows: int) -> List[List[Any]]:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows > 0 and i >= max_rows + 1:
                break
            data.append([str(cell) if cell is not None else '' for cell in row])

        wb.close()
        return data

    def get_required_params(self) -> List[str]:
        return ['file_path']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'sheet_name': '', 'output_var': 'spreadsheet_data', 'max_rows': 0}


class SpreadsheetWriteAction(BaseAction):
    """Write data to spreadsheet."""
    action_type = "spreadsheet_write"
    display_name = "写入电子表格"
    description = "将数据写入Excel或CSV文件"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute write.

        Args:
            context: Execution context.
            params: Dict with file_path, data, sheet_name.

        Returns:
            ActionResult indicating success.
        """
        file_path = params.get('file_path', '')
        data_param = params.get('data', [])
        sheet_name = params.get('sheet_name', 'Sheet1')

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_data = context.resolve_value(data_param)
            resolved_sheet = context.resolve_value(sheet_name)

            if not isinstance(resolved_data, list):
                return ActionResult(
                    success=False,
                    message="data参数必须是列表"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                self._write_csv(resolved_path, resolved_data)
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                self._write_excel(resolved_path, resolved_data, resolved_sheet)
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            return ActionResult(
                success=True,
                message=f"已写入 {len(resolved_data)} 行到 {resolved_path}",
                data={'rows': len(resolved_data), 'path': resolved_path}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"写入电子表格失败: {str(e)}"
            )

    def _write_csv(self, path: str, data: List[List]) -> None:
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(data)

    def _write_excel(self, path: str, data: List[List], sheet_name: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for row in data:
            ws.append(row)

        wb.save(path)
        wb.close()

    def get_required_params(self) -> List[str]:
        return ['file_path', 'data']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'sheet_name': 'Sheet1'}


class SpreadsheetCreateAction(BaseAction):
    """Create new spreadsheet."""
    action_type = "spreadsheet_create"
    display_name = "创建电子表格"
    description = "创建新的Excel或CSV文件"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute create.

        Args:
            context: Execution context.
            params: Dict with file_path, headers.

        Returns:
            ActionResult indicating success.
        """
        file_path = params.get('file_path', '')
        headers = params.get('headers', [])

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_headers = context.resolve_value(headers)

            if os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件已存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                if resolved_headers:
                    with open(resolved_path, 'w', encoding='utf-8-sig', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(resolved_headers)
                else:
                    open(resolved_path, 'w', encoding='utf-8-sig').close()
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                wb = Workbook()
                ws = wb.active
                if resolved_headers:
                    ws.append(resolved_headers)
                wb.save(resolved_path)
                wb.close()
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            return ActionResult(
                success=True,
                message=f"已创建电子表格: {resolved_path}",
                data={'path': resolved_path}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"创建电子表格失败: {str(e)}"
            )

    def get_required_params(self) -> List[str]:
        return ['file_path']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'headers': []}


class SpreadsheetCellReadAction(BaseAction):
    """Read specific cell."""
    action_type = "spreadsheet_cell_read"
    display_name = "读取单元格"
    description = "读取电子表格指定单元格的值"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute cell read.

        Args:
            context: Execution context.
            params: Dict with file_path, cell, sheet_name, output_var.

        Returns:
            ActionResult with cell value.
        """
        file_path = params.get('file_path', '')
        cell = params.get('cell', 'A1')
        sheet_name = params.get('sheet_name', '')
        output_var = params.get('output_var', 'cell_value')

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        valid, msg = self.validate_type(cell, str, 'cell')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_cell = context.resolve_value(cell)

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                value = self._read_csv_cell(resolved_path, resolved_cell)
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                sheet = context.resolve_value(sheet_name) if sheet_name else ''
                value = self._read_excel_cell(resolved_path, resolved_cell, sheet)
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            context.set(output_var, value)

            return ActionResult(
                success=True,
                message=f"单元格 {resolved_cell}: {value}",
                data={'cell': resolved_cell, 'value': value, 'output_var': output_var}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"读取单元格失败: {str(e)}"
            )

    def _read_csv_cell(self, path: str, cell: str) -> str:
        col_letter = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell))) - 1
        col_num = 0
        for i, c in enumerate(reversed(col_letter)):
            col_num += (ord(c.upper()) - ord('A') + 1) * (26 ** i)

        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = list(csv.reader(f))
            if row_num < len(reader):
                row = reader[row_num]
                if col_num - 1 < len(row):
                    return str(row[col_num - 1])
        return ''

    def _read_excel_cell(self, path: str, cell: str, sheet_name: str) -> Any:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        value = ws[cell].value
        wb.close()
        return value

    def get_required_params(self) -> List[str]:
        return ['file_path', 'cell']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'sheet_name': '', 'output_var': 'cell_value'}


class SpreadsheetCellWriteAction(BaseAction):
    """Write to specific cell."""
    action_type = "spreadsheet_cell_write"
    display_name = "写入单元格"
    description = "向电子表格指定单元格写入值"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute cell write.

        Args:
            context: Execution context.
            params: Dict with file_path, cell, value, sheet_name.

        Returns:
            ActionResult indicating success.
        """
        file_path = params.get('file_path', '')
        cell = params.get('cell', 'A1')
        value = params.get('value', '')
        sheet_name = params.get('sheet_name', '')

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        valid, msg = self.validate_type(cell, str, 'cell')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_cell = context.resolve_value(cell)
            resolved_value = context.resolve_value(value)

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                self._write_csv_cell(resolved_path, resolved_cell, str(resolved_value))
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                sheet = context.resolve_value(sheet_name) if sheet_name else ''
                self._write_excel_cell(resolved_path, resolved_cell, resolved_value, sheet)
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            return ActionResult(
                success=True,
                message=f"已写入 {resolved_cell} = {resolved_value}",
                data={'cell': resolved_cell, 'value': resolved_value}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"写入单元格失败: {str(e)}"
            )

    def _write_csv_cell(self, path: str, cell: str, value: str) -> None:
        col_letter = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell))) - 1
        col_num = 0
        for i, c in enumerate(reversed(col_letter)):
            col_num += (ord(c.upper()) - ord('A') + 1) * (26 ** i)

        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = list(csv.reader(f))

        # Extend rows if needed
        while len(reader) <= row_num:
            reader.append([])

        # Extend columns if needed
        while len(reader[row_num]) < col_num:
            reader[row_num].append('')

        reader[row_num][col_num - 1] = value

        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(reader)

    def _write_excel_cell(self, path: str, cell: str, value: Any, sheet_name: str) -> None:
        wb = openpyxl.load_workbook(path)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        ws[cell] = value
        wb.save(path)
        wb.close()

    def get_required_params(self) -> List[str]:
        return ['file_path', 'cell', 'value']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'sheet_name': ''}


class SpreadsheetAppendRowAction(BaseAction):
    """Append row of data."""
    action_type = "spreadsheet_append_row"
    display_name = "追加行数据"
    description = "向电子表格追加一行数据"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute append row.

        Args:
            context: Execution context.
            params: Dict with file_path, row, sheet_name.

        Returns:
            ActionResult indicating success.
        """
        file_path = params.get('file_path', '')
        row = params.get('row', [])
        sheet_name = params.get('sheet_name', '')

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_row = context.resolve_value(row)

            if not isinstance(resolved_row, list):
                return ActionResult(
                    success=False,
                    message="row参数必须是列表"
                )

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                with open(resolved_path, 'a', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(resolved_row)
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                wb = openpyxl.load_workbook(resolved_path)
                if sheet_name:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                ws.append(resolved_row)
                wb.save(resolved_path)
                wb.close()
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            return ActionResult(
                success=True,
                message=f"已追加行: {resolved_row}",
                data={'row': resolved_row}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"追加行失败: {str(e)}"
            )

    def get_required_params(self) -> List[str]:
        return ['file_path', 'row']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'sheet_name': ''}


class SpreadsheetFilterAction(BaseAction):
    """Filter rows by condition."""
    action_type = "spreadsheet_filter"
    display_name = "筛选电子表格"
    description = "根据条件筛选电子表格行"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute filter.

        Args:
            context: Execution context.
            params: Dict with file_path, column, operator, value, output_var.

        Returns:
            ActionResult with filtered rows.
        """
        file_path = params.get('file_path', '')
        column = params.get('column', 0)
        operator = params.get('operator', 'equals')
        value = params.get('value', '')
        output_var = params.get('output_var', 'filtered_rows')
        has_header = params.get('has_header', True)

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_col = context.resolve_value(column)
            resolved_op = context.resolve_value(operator)
            resolved_val = context.resolve_value(value)
            resolved_header = context.resolve_value(has_header)

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                with open(resolved_path, 'r', encoding='utf-8-sig') as f:
                    all_rows = list(csv.reader(f))
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                wb = openpyxl.load_workbook(resolved_path, data_only=True)
                ws = wb.active
                all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
                wb.close()
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            if not all_rows:
                return ActionResult(
                    success=False,
                    message="文件为空"
                )

            header = all_rows[0] if resolved_header else []
            data_rows = all_rows[1:] if resolved_header else all_rows

            col_idx = resolved_col
            if isinstance(resolved_col, str):
                col_idx = 0
                for i, h in enumerate(header):
                    if str(h).lower() == resolved_col.lower():
                        col_idx = i
                        break

            filtered = []
            for row in data_rows:
                if col_idx < len(row):
                    cell_val = str(row[col_idx])
                    if self._match(cell_val, resolved_op, str(resolved_val)):
                        filtered.append(row)

            context.set(output_var, filtered)

            return ActionResult(
                success=True,
                message=f"筛选结果: {len(filtered)} 行",
                data={'rows': len(filtered), 'output_var': output_var}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"筛选失败: {str(e)}"
            )

    def _match(self, cell_val: str, op: str, target: str) -> bool:
        if op == 'equals':
            return cell_val == target
        elif op == 'contains':
            return target in cell_val
        elif op == 'starts_with':
            return cell_val.startswith(target)
        elif op == 'ends_with':
            return cell_val.endswith(target)
        elif op == 'greater':
            try:
                return float(cell_val) > float(target)
            except ValueError:
                return cell_val > target
        elif op == 'less':
            try:
                return float(cell_val) < float(target)
            except ValueError:
                return cell_val < target
        elif op == 'not_equals':
            return cell_val != target
        return False

    def get_required_params(self) -> List[str]:
        return ['file_path', 'column', 'operator', 'value']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'output_var': 'filtered_rows', 'has_header': True}


class SpreadsheetSortAction(BaseAction):
    """Sort spreadsheet."""
    action_type = "spreadsheet_sort"
    display_name = "排序电子表格"
    description = "对电子表格按列排序"
    version = "1.0"

    def execute(
        self,
        context: Any,
        params: Dict[str, Any]
    ) -> ActionResult:
        """Execute sort.

        Args:
            context: Execution context.
            params: Dict with file_path, column, order, output_var.

        Returns:
            ActionResult indicating success.
        """
        file_path = params.get('file_path', '')
        column = params.get('column', 0)
        order = params.get('order', 'asc')
        output_var = params.get('output_var', 'sorted_data')
        has_header = params.get('has_header', True)

        valid, msg = self.validate_type(file_path, str, 'file_path')
        if not valid:
            return ActionResult(success=False, message=msg)

        try:
            resolved_path = context.resolve_value(file_path)
            resolved_col = context.resolve_value(column)
            resolved_order = context.resolve_value(order)
            resolved_header = context.resolve_value(has_header)

            if not os.path.exists(resolved_path):
                return ActionResult(
                    success=False,
                    message=f"文件不存在: {resolved_path}"
                )

            ext = os.path.splitext(resolved_path)[1].lower()

            if ext == '.csv':
                with open(resolved_path, 'r', encoding='utf-8-sig') as f:
                    all_rows = list(csv.reader(f))
            elif ext in ('.xlsx', '.xls'):
                if not OPENPYXL_AVAILABLE:
                    return ActionResult(
                        success=False,
                        message="openpyxl未安装"
                    )
                wb = openpyxl.load_workbook(resolved_path, data_only=True)
                ws = wb.active
                all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
                wb.close()
            else:
                return ActionResult(
                    success=False,
                    message=f"不支持的文件格式: {ext}"
                )

            if not all_rows:
                return ActionResult(
                    success=False,
                    message="文件为空"
                )

            header = all_rows[0] if resolved_header else []
            data_rows = all_rows[1:] if resolved_header else all_rows

            col_idx = resolved_col
            reverse = resolved_order == 'desc'

            sorted_rows = sorted(
                data_rows,
                key=lambda r: self._sort_key(r[col_idx] if col_idx < len(r) else ''),
                reverse=reverse
            )

            if resolved_header:
                sorted_rows = [header] + sorted_rows

            context.set(output_var, sorted_rows)

            return ActionResult(
                success=True,
                message=f"已排序 {len(data_rows)} 行",
                data={'rows': len(sorted_rows), 'output_var': output_var}
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"排序失败: {str(e)}"
            )

    def _sort_key(self, val: Any) -> Any:
        if val is None:
            return ''
        try:
            return float(val)
        except (ValueError, TypeError):
            return str(val)

    def get_required_params(self) -> List[str]:
        return ['file_path', 'column']

    def get_optional_params(self) -> Dict[str, Any]:
        return {'order': 'asc', 'output_var': 'sorted_data', 'has_header': True}
