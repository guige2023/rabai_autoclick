"""
Data Exporter Action Module.

Provides multi-format data export capabilities including
CSV, Excel, JSON, XML, Parquet, and custom formats.

Author: rabai_autoclick team
"""

import io
import csv
import logging
from typing import (
    Optional, Dict, Any, List, Union, Callable, Type
)
from dataclasses import dataclass, field, is_dataclass, asdict
from enum import Enum
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Supported export formats."""
    CSV = "csv"
    TSV = "tsv"
    JSON = "json"
    JSONL = "jsonl"
    XML = "xml"
    EXCEL = "xlsx"
    PARQUET = "parquet"
    HTML = "html"
    Markdown = "md"
    SQL = "sql"
    CUSTOM = "custom"


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    format: ExportFormat = ExportFormat.CSV
    delimiter: str = ","
    include_header: bool = True
    encoding: str = "utf-8"
    datetime_format: str = "%Y-%m-%d %H:%M:%S"
    null_value: str = ""
    precision: Optional[int] = None
    max_rows: Optional[int] = None
    sheet_name: str = "Sheet1"
    pretty_print: bool = True
    custom_formatter: Optional[Callable] = None


@dataclass
class ExportResult:
    """Result of an export operation."""
    success: bool
    data: Optional[Union[str, bytes]] = None
    file_path: Optional[Path] = None
    row_count: int = 0
    error: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


class DataExporterAction:
    """
    Multi-Format Data Exporter.

    Supports exporting data to various formats with
    customizable configuration and streaming support.

    Example:
        >>> exporter = DataExporterAction()
        >>> data = [{"name": "John", "age": 30}, {"name": "Jane", "age": 25}]
        >>> result = exporter.export(data, format=ExportFormat.CSV)
    """

    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()
        self._formatters: Dict[ExportFormat, Callable] = {}

    def export(
        self,
        data: List[Dict[str, Any]],
        format: Optional[ExportFormat] = None,
        file_path: Optional[Union[str, Path]] = None,
        config: Optional[ExportConfig] = None,
    ) -> ExportResult:
        """
        Export data to specified format.

        Args:
            data: List of data records
            format: Export format (uses config default if None)
            file_path: Optional file path to write to
            config: Optional export configuration

        Returns:
            ExportResult with exported data or file path
        """
        cfg = config or self.config
        fmt = format or cfg.format

        if cfg.max_rows and len(data) > cfg.max_rows:
            data = data[: cfg.max_rows]

        try:
            if fmt == ExportFormat.CSV:
                result_data = self._export_csv(data, cfg)
            elif fmt == ExportFormat.TSV:
                result_data = self._export_tsv(data, cfg)
            elif fmt == ExportFormat.JSON:
                result_data = self._export_json(data, cfg)
            elif fmt == ExportFormat.JSONL:
                result_data = self._export_jsonl(data, cfg)
            elif fmt == ExportFormat.XML:
                result_data = self._export_xml(data, cfg)
            elif fmt == ExportFormat.EXCEL:
                result_data = self._export_excel(data, cfg)
            elif fmt == ExportFormat.PARQUET:
                result_data = self._export_parquet(data, cfg)
            elif fmt == ExportFormat.HTML:
                result_data = self._export_html(data, cfg)
            elif fmt == ExportFormat.MARKDOWN:
                result_data = self._export_markdown(data, cfg)
            elif fmt == ExportFormat.SQL:
                result_data = self._export_sql(data, cfg)
            elif fmt == ExportFormat.CUSTOM and cfg.custom_formatter:
                result_data = cfg.custom_formatter(data)
            else:
                raise ValueError(f"Unsupported format: {fmt}")

            result = ExportResult(
                success=True,
                data=result_data,
                row_count=len(data),
                metadata={"format": fmt.value},
            )

            if file_path:
                self._write_file(result_data, Path(file_path), cfg)
                result.file_path = Path(file_path)

            return result

        except Exception as e:
            logger.error(f"Export failed: {e}")
            return ExportResult(success=False, error=str(e))

    def _format_value(self, value: Any, cfg: ExportConfig) -> str:
        """Format a single value for export."""
        if value is None:
            return cfg.null_value

        if isinstance(value, datetime):
            return value.strftime(cfg.datetime_format)

        if isinstance(value, float) and cfg.precision is not None:
            return f"{value:.{cfg.precision}f}"

        if isinstance(value, (dict, list)):
            return str(value)

        return str(value)

    def _export_csv(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to CSV format."""
        if not data:
            return ""

        output = io.StringIO()
        fieldnames = list(data[0].keys())

        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=cfg.delimiter,
            extrasaction="ignore",
        )

        if cfg.include_header:
            writer.writeheader()

        for row in data:
            formatted_row = {k: self._format_value(v, cfg) for k, v in row.items()}
            writer.writerow(formatted_row)

        return output.getvalue()

    def _export_tsv(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to TSV format."""
        cfg_delimiter = cfg.delimiter
        cfg.delimiter = "\t"
        result = self._export_csv(data, cfg)
        cfg.delimiter = cfg_delimiter
        return result

    def _export_json(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to JSON format."""
        import json

        serialized = [self._serialize_record(r, cfg) for r in data]

        if cfg.pretty_print:
            return json.dumps(serialized, indent=2, ensure_ascii=False)
        return json.dumps(serialized, ensure_ascii=False)

    def _serialize_record(self, record: Dict, cfg: ExportConfig) -> Dict:
        """Serialize a record for JSON export."""
        result = {}
        for key, value in record.items():
            if value is None:
                result[key] = None
            elif isinstance(value, datetime):
                result[key] = value.strftime(cfg.datetime_format)
            elif isinstance(value, float) and cfg.precision is not None:
                result[key] = round(value, cfg.precision)
            else:
                result[key] = value
        return result

    def _export_jsonl(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to JSONL format."""
        import json
        lines = []
        for record in data:
            serialized = self._serialize_record(record, cfg)
            lines.append(json.dumps(serialized, ensure_ascii=False))
        return "\n".join(lines)

    def _export_xml(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to XML format."""
        import xml.etree.ElementTree as ET

        root = ET.Element("data")
        for record in data:
            item = ET.SubElement(root, "record")
            for key, value in record.items():
                child = ET.SubElement(item, str(key))
                child.text = self._format_value(value, cfg)
                if value is None:
                    child.set("nil", "true")

        return ET.tostring(root, encoding="unicode")

    def _export_excel(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> bytes:
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = cfg.sheet_name

            if not data:
                return b""

            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row_idx, record in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            logger.warning("openpyxl not available, using CSV")
            csv_data = self._export_csv(data, cfg)
            return csv_data.encode(cfg.encoding)

    def _export_parquet(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> bytes:
        """Export to Parquet format."""
        try:
            import pandas as pd
            import pyarrow.parquet as pq

            df = pd.DataFrame(data)
            buffer = io.BytesIO()
            df.to_parquet(buffer, engine="pyarrow", index=False)
            return buffer.getvalue()

        except ImportError:
            logger.warning("pandas/pyarrow not available, using CSV")
            csv_data = self._export_csv(data, cfg)
            return csv_data.encode(cfg.encoding)

    def _export_html(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to HTML format."""
        if not data:
            return "<table></table>"

        headers = list(data[0].keys())
        html = ['<table border="1">']

        if cfg.include_header:
            html.append("<thead><tr>")
            for header in headers:
                html.append(f"<th>{self._escape_html(header)}</th>")
            html.append("</tr></thead>")

        html.append("<tbody>")
        for record in data:
            html.append("<tr>")
            for header in headers:
                value = record.get(header)
                html.append(f"<td>{self._escape_html(self._format_value(value, cfg))}</td>")
            html.append("</tr>")
        html.append("</tbody></table>")

        return "".join(html)

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;")
        )

    def _export_markdown(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to Markdown table format."""
        if not data:
            return ""

        headers = list(data[0].keys())
        lines = []

        header_line = "| " + " | ".join(headers) + " |"
        lines.append(header_line)

        separator = "| " + " | ".join(["---"] * len(headers)) + " |"
        lines.append(separator)

        for record in data:
            row = "| " + " | ".join(
                self._format_value(record.get(h), cfg) for h in headers
            ) + " |"
            lines.append(row)

        return "\n".join(lines)

    def _export_sql(self, data: List[Dict[str, Any]], cfg: ExportConfig) -> str:
        """Export to SQL INSERT statements."""
        if not data:
            return ""

        table_name = cfg.metadata.get("table_name", "data")
        columns = list(data[0].keys())

        lines = []
        for record in data:
            values = []
            for value in (record.get(col) for col in columns):
                if value is None:
                    values.append("NULL")
                elif isinstance(value, (int, float)):
                    values.append(str(value))
                else:
                    escaped = str(value).replace("'", "''")
                    values.append(f"'{escaped}'")

            line = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(values)});"
            lines.append(line)

        return "\n".join(lines)

    def _write_file(
        self,
        data: Union[str, bytes],
        file_path: Path,
        cfg: ExportConfig,
    ) -> None:
        """Write data to file."""
        file_path.parent.mkdir(parents=True, exist_ok=True)

        if isinstance(data, str):
            file_path.write_text(data, encoding=cfg.encoding)
        else:
            file_path.write_bytes(data)

    def export_to_file(
        self,
        data: List[Dict[str, Any]],
        file_path: Union[str, Path],
        format: Optional[ExportFormat] = None,
    ) -> ExportResult:
        """
        Export data directly to file.

        Args:
            data: Data records
            file_path: Output file path
            format: Export format (inferred from extension if None)

        Returns:
            ExportResult
        """
        path = Path(file_path)

        if format is None:
            ext = path.suffix.lower().lstrip(".")
            format_map = {
                "csv": ExportFormat.CSV,
                "tsv": ExportFormat.TSV,
                "json": ExportFormat.JSON,
                "jsonl": ExportFormat.JSONL,
                "xml": ExportFormat.XML,
                "xlsx": ExportFormat.EXCEL,
                "html": ExportFormat.HTML,
                "md": ExportFormat.MARKDOWN,
                "parquet": ExportFormat.PARQUET,
            }
            format = format_map.get(ext, ExportFormat.CSV)

        return self.export(data, format=format, file_path=path)

    def export_streaming(
        self,
        data: List[Dict[str, Any]],
        format: ExportFormat,
        chunk_size: int = 1000,
    ) -> List[Union[str, bytes]]:
        """
        Export data in chunks for streaming.

        Args:
            data: Data records
            format: Export format
            chunk_size: Records per chunk

        Returns:
            List of data chunks
        """
        chunks = []
        for i in range(0, len(data), chunk_size):
            chunk = data[i : i + chunk_size]
            result = self.export(chunk, format=format)
            if result.success and result.data:
                chunks.append(result.data)
        return chunks
