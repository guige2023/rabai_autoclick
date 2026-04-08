"""Excel action module for RabAI AutoClick.

Provides Excel file operations including reading, writing, cell manipulation,
formula support, and format preservation using openpyxl.
"""

import os
import sys
import time
from typing import Any, Dict, List, Optional, Union, Tuple
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.base_action import BaseAction, ActionResult


class ExcelWorkbook:
    """Excel workbook wrapper with read/write operations.
    
    Provides a high-level interface for Excel file operations
    including cell access, sheet management, and format preservation.
    """
    
    def __init__(self, path: Optional[str] = None) -> None:
        """Initialize Excel workbook.
        
        Args:
            path: Optional path to an existing Excel file.
        """
        self._path = path
        self._workbook = None
        self._loaded = False
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel support. Install with: pip install openpyxl"
            )
        
        if path and os.path.exists(path):
            self._load(path)
        else:
            self._workbook = Workbook()
            self._loaded = True
    
    def _load(self, path: str) -> None:
        """Load an existing Excel file.
        
        Args:
            path: Path to the Excel file.
        """
        from openpyxl import load_workbook
        self._workbook = load_workbook(path)
        self._path = path
        self._loaded = True
    
    def save(self, path: Optional[str] = None) -> None:
        """Save the workbook to a file.
        
        Args:
            path: Optional path to save to. Uses original path if not provided.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        save_path = path or self._path
        if not save_path:
            raise ValueError("No save path specified")
        
        self._workbook.save(save_path)
        self._path = save_path
    
    def get_sheet(self, name: Optional[str] = None, index: int = 0) -> "ExcelSheet":
        """Get a sheet by name or index.
        
        Args:
            name: Sheet name to retrieve.
            index: Sheet index (0-based) if name not provided.
            
        Returns:
            ExcelSheet wrapper for the requested sheet.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        if name:
            ws = self._workbook[name]
        else:
            sheet_names = self._workbook.sheetnames
            if index < len(sheet_names):
                ws = self._workbook[sheet_names[index]]
            else:
                raise IndexError(f"Sheet index {index} out of range")
        
        return ExcelSheet(ws)
    
    def create_sheet(
        self,
        name: str,
        index: Optional[int] = None
    ) -> "ExcelSheet":
        """Create a new sheet.
        
        Args:
            name: Name for the new sheet.
            index: Optional position to insert the sheet.
            
        Returns:
            ExcelSheet wrapper for the new sheet.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        ws = self._workbook.create_sheet(title=name, index=index)
        return ExcelSheet(ws)
    
    def delete_sheet(self, name: str) -> bool:
        """Delete a sheet by name.
        
        Args:
            name: Name of the sheet to delete.
            
        Returns:
            True if deleted, False if sheet didn't exist.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        if name in self._workbook.sheetnames:
            del self._workbook[name]
            return True
        return False
    
    def list_sheets(self) -> List[str]:
        """List all sheet names in the workbook.
        
        Returns:
            List of sheet names.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        return self._workbook.sheetnames
    
    @property
    def active_sheet(self) -> "ExcelSheet":
        """Get the currently active sheet.
        
        Returns:
            ExcelSheet wrapper for the active sheet.
        """
        if not self._loaded:
            raise RuntimeError("Workbook not loaded or created")
        
        return ExcelSheet(self._workbook.active)


class ExcelSheet:
    """Excel sheet wrapper with cell-level operations.
    
    Provides methods for reading, writing, and formatting cells.
    """
    
    def __init__(self, worksheet: Any) -> None:
        """Initialize Excel sheet wrapper.
        
        Args:
            worksheet: openpyxl worksheet object.
        """
        self._ws = worksheet
    
    @property
    def name(self) -> str:
        """Get the sheet name."""
        return self._ws.title
    
    @property
    def dimensions(self) -> str:
        """Get the sheet dimensions as a string (e.g., 'A1:E10')."""
        return self._ws.dimensions
    
    def get_cell(
        self,
        row: int,
        col: int
    ) -> Optional[Any]:
        """Get the value of a cell.
        
        Args:
            row: Row number (1-based).
            col: Column number (1-based).
            
        Returns:
            Cell value or None if cell is empty.
        """
        return self._ws.cell(row=row, column=col).value
    
    def set_cell(
        self,
        row: int,
        col: int,
        value: Any,
        number_format: Optional[str] = None
    ) -> None:
        """Set the value and optionally the format of a cell.
        
        Args:
            row: Row number (1-based).
            col: Column number (1-based).
            value: Value to set.
            number_format: Optional Excel number format string.
        """
        cell = self._ws.cell(row=row, column=col)
        cell.value = value
        
        if number_format:
            cell.number_format = number_format
    
    def read_range(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int
    ) -> List[List[Any]]:
        """Read a rectangular range of cells.
        
        Args:
            start_row: Starting row (1-based).
            start_col: Starting column (1-based).
            end_row: Ending row (1-based).
            end_col: Ending column (1-based).
            
        Returns:
            2D list of cell values.
        """
        result: List[List[Any]] = []
        
        for row in range(start_row, end_row + 1):
            row_data: List[Any] = []
            for col in range(start_col, end_col + 1):
                row_data.append(self._ws.cell(row=row, column=col).value)
            result.append(row_data)
        
        return result
    
    def write_range(
        self,
        start_row: int,
        start_col: int,
        data: List[List[Any]]
    ) -> None:
        """Write a 2D list of values to a range.
        
        Args:
            start_row: Starting row (1-based).
            start_col: Starting column (1-based).
            data: 2D list of values to write.
        """
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                self._ws.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx
                ).value = value
    
    def set_header(
        self,
        headers: List[str],
        row: int = 1,
        col: int = 1,
        bold: bool = True,
        bg_color: Optional[str] = None
    ) -> None:
        """Set header row with optional formatting.
        
        Args:
            headers: List of header names.
            row: Row number for the header (1-based).
            col: Starting column number (1-based).
            bold: Whether to make headers bold.
            bg_color: Optional background color (hex string).
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        for col_idx, header in enumerate(headers):
            cell = self._ws.cell(row=row, column=col + col_idx)
            cell.value = header
            
            if bold:
                cell.font = Font(bold=True)
            
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def auto_fit_columns(self, min_width: float = 5.0, max_width: float = 50.0) -> None:
        """Auto-fit column widths based on content.
        
        Args:
            min_width: Minimum column width.
            max_width: Maximum column width.
        """
        for column in self._ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            self._ws.column_dimensions[column_letter].width = adjusted_width
    
    def clear(self) -> None:
        """Clear all cell values and formatting from the sheet."""
        for row in self._ws.iter_rows():
            for cell in row:
                cell.value = None


class ExcelAction(BaseAction):
    """Excel action for reading and writing Excel files.
    
    Supports reading, writing, cell manipulation, and formatting.
    """
    action_type: str = "excel"
    display_name: str = "Excel动作"
    description: str = "读取和写入Excel文件，支持单元格操作和格式化"
    
    def get_required_params(self) -> List[str]:
        """Return required parameters for this action."""
        return ["operation"]
    
    def execute(self, context: Any, params: Dict[str, Any]) -> ActionResult:
        """Execute Excel operation.
        
        Args:
            context: Execution context.
            params: Operation and parameters.
            
        Returns:
            ActionResult with operation outcome.
        """
        start_time = time.time()
        
        try:
            operation = params.get("operation", "read")
            
            if operation == "read":
                return self._read_excel(params, start_time)
            elif operation == "write":
                return self._write_excel(params, start_time)
            elif operation == "get_cell":
                return self._get_cell(params, start_time)
            elif operation == "set_cell":
                return self._set_cell(params, start_time)
            elif operation == "list_sheets":
                return self._list_sheets(params, start_time)
            elif operation == "create_sheet":
                return self._create_sheet(params, start_time)
            elif operation == "delete_sheet":
                return self._delete_sheet(params, start_time)
            else:
                return ActionResult(
                    success=False,
                    message=f"Unknown operation: {operation}",
                    duration=time.time() - start_time
                )
        
        except ImportError as e:
            return ActionResult(
                success=False,
                message=f"Import error: {str(e)}",
                duration=time.time() - start_time
            )
        except Exception as e:
            return ActionResult(
                success=False,
                message=f"Excel operation failed: {str(e)}",
                duration=time.time() - start_time
            )
    
    def _read_excel(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Read data from an Excel file."""
        path = params.get("path", "")
        
        if not path:
            return ActionResult(
                success=False,
                message="Path is required for read operation",
                duration=time.time() - start_time
            )
        
        if not os.path.exists(path):
            return ActionResult(
                success=False,
                message=f"Excel file not found: {path}",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook(path)
        sheet_name = params.get("sheet")
        sheet_index = params.get("sheet_index", 0)
        
        try:
            if sheet_name:
                sheet = wb.get_sheet(name=sheet_name)
            else:
                sheet = wb.get_sheet(index=sheet_index)
        except (IndexError, KeyError):
            return ActionResult(
                success=False,
                message=f"Sheet not found: {sheet_name or f'index {sheet_index}'}",
                duration=time.time() - start_time
            )
        
        data = sheet.read_range(
            start_row=params.get("start_row", 1),
            start_col=params.get("start_col", 1),
            end_row=params.get("end_row", sheet._ws.max_row),
            end_col=params.get("end_col", sheet._ws.max_column)
        )
        
        return ActionResult(
            success=True,
            message=f"Read {len(data)} rows from sheet: {sheet.name}",
            data={"sheet": sheet.name, "data": data},
            duration=time.time() - start_time
        )
    
    def _write_excel(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Write data to an Excel file."""
        path = params.get("path", "")
        data = params.get("data", [])
        sheet_name = params.get("sheet", "Sheet1")
        
        if not path:
            return ActionResult(
                success=False,
                message="Path is required for write operation",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook()
        
        if os.path.exists(path):
            try:
                wb = ExcelWorkbook(path)
            except Exception:
                wb = ExcelWorkbook()
        
        try:
            sheet = wb.get_sheet(name=sheet_name)
        except KeyError:
            sheet = wb.create_sheet(name=sheet_name)
        
        start_row = params.get("start_row", 1)
        start_col = params.get("start_col", 1)
        
        sheet.write_range(start_row, start_col, data)
        
        if params.get("auto_fit", False):
            sheet.auto_fit_columns()
        
        if params.get("set_header") and data:
            headers = params.get("header_row", 1)
            if isinstance(headers, list):
                sheet.set_header(headers, row=start_row, col=start_col)
        
        wb.save(path)
        
        return ActionResult(
            success=True,
            message=f"Wrote {len(data)} rows to {path}",
            data={"rows_written": len(data)},
            duration=time.time() - start_time
        )
    
    def _get_cell(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Get a single cell value."""
        path = params.get("path", "")
        row = params.get("row", 1)
        col = params.get("col", 1)
        sheet_name = params.get("sheet")
        
        if not path:
            return ActionResult(
                success=False,
                message="Path is required",
                duration=time.time() - start_time
            )
        
        if not os.path.exists(path):
            return ActionResult(
                success=False,
                message=f"Excel file not found: {path}",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook(path)
        
        try:
            if sheet_name:
                sheet = wb.get_sheet(name=sheet_name)
            else:
                sheet = wb.active_sheet
        except (IndexError, KeyError):
            return ActionResult(
                success=False,
                message=f"Sheet not found: {sheet_name}",
                duration=time.time() - start_time
            )
        
        value = sheet.get_cell(row, col)
        
        return ActionResult(
            success=True,
            message=f"Retrieved cell ({row}, {col}): {value}",
            data={"row": row, "col": col, "value": value},
            duration=time.time() - start_time
        )
    
    def _set_cell(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Set a single cell value."""
        path = params.get("path", "")
        row = params.get("row", 1)
        col = params.get("col", 1)
        value = params.get("value")
        sheet_name = params.get("sheet")
        number_format = params.get("number_format")
        
        if not path:
            return ActionResult(
                success=False,
                message="Path is required",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook() if not os.path.exists(path) else ExcelWorkbook(path)
        
        try:
            if sheet_name:
                sheet = wb.get_sheet(name=sheet_name)
            else:
                sheet = wb.active_sheet
        except KeyError:
            sheet = wb.create_sheet(name=sheet_name or "Sheet1")
        
        sheet.set_cell(row, col, value, number_format)
        wb.save(path)
        
        return ActionResult(
            success=True,
            message=f"Set cell ({row}, {col}) = {value}",
            duration=time.time() - start_time
        )
    
    def _list_sheets(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """List all sheets in an Excel file."""
        path = params.get("path", "")
        
        if not path or not os.path.exists(path):
            return ActionResult(
                success=False,
                message=f"Excel file not found: {path}",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook(path)
        sheets = wb.list_sheets()
        
        return ActionResult(
            success=True,
            message=f"Found {len(sheets)} sheets",
            data={"sheets": sheets},
            duration=time.time() - start_time
        )
    
    def _create_sheet(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Create a new sheet in an Excel file."""
        path = params.get("path", "")
        sheet_name = params.get("sheet_name", "NewSheet")
        
        if not path:
            return ActionResult(
                success=False,
                message="Path is required",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook() if not os.path.exists(path) else ExcelWorkbook(path)
        sheet = wb.create_sheet(sheet_name)
        wb.save(path)
        
        return ActionResult(
            success=True,
            message=f"Created sheet: {sheet_name}",
            duration=time.time() - start_time
        )
    
    def _delete_sheet(self, params: Dict[str, Any], start_time: float) -> ActionResult:
        """Delete a sheet from an Excel file."""
        path = params.get("path", "")
        sheet_name = params.get("sheet_name", "")
        
        if not path or not os.path.exists(path):
            return ActionResult(
                success=False,
                message=f"Excel file not found: {path}",
                duration=time.time() - start_time
            )
        
        if not sheet_name:
            return ActionResult(
                success=False,
                message="sheet_name is required",
                duration=time.time() - start_time
            )
        
        wb = ExcelWorkbook(path)
        deleted = wb.delete_sheet(sheet_name)
        
        if deleted:
            wb.save(path)
        
        return ActionResult(
            success=deleted,
            message=f"Deleted sheet: {sheet_name}" if deleted else f"Sheet not found: {sheet_name}",
            duration=time.time() - start_time
        )
